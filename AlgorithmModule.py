
import cv2
#from PyQt5.QtWidgets import QDialog, QApplication
#from PyQt5 import QtGui, QtWidgets, Qt, QtCore, QtSql
import sys
#from gui import Ui_MainWindow
#import serial.tools.list_ports
#import time
#from PyQt5.QtCore import QIODevice
import os

import keras.preprocessing.image


#import pandas as pd
import tensorflow as tf
from openpyxl import load_workbook
import random
import numpy as np

##
"""
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import backend as K
from tensorflow.keras.layers import Dense, Activation,Dropout,Conv2D, MaxPooling2D,BatchNormalization, Flatten
from tensorflow.keras.optimizers import Adam, Adamax
from tensorflow.keras.metrics import categorical_crossentropy
from tensorflow.keras import regularizers
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.keras.models import Model, load_model, Sequential
import numpy as np
import pandas as pd
import shutil
import time
import cv2 as cv2
from tqdm import tqdm
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from matplotlib.pyplot import imshow
import seaborn as sns
sns.set_style('darkgrid')
from PIL import Image
from sklearn.metrics import confusion_matrix, classification_report
from IPython.core.display import display, HTML
# stop annoying tensorflow warning messages
import logging
logging.getLogger("tensorflow").setLevel(logging.ERROR)
print ('modules loaded')
"""
##

class Algorithm():

	def __init__(self):
		self.model_file='EfficientNetB3-fruits-100.0.h5'
		self.class_file='classes.xlsx'
		self.classesNames = self.readClasses(self.class_file)
		self.classesNum = len(self.classesNames)
		self.model = self.initCnn(os.getcwd() + "/ML_Models/"+self.model_file)

		print("2: classes are uploaded: ")
		#for i in range(1,len(self.classesNames)):
		#	print("   ", self.classesNames[i], end="\n")
		print(self.classesNames)

	def readClasses(self, file):
		global classesDict
		classesDict={}
		wb=load_workbook(file)
		ws=wb.get_sheet_by_name("Sheet1")

		column = ws['A']
		t = len(column)
		idCol=[column[x].value for x in range( t)]
		column = ws['B']
		classCol = [column[x].value for x in range(t)]
		column = ws['C']
		cpriceCol = [column[x].value for x in range(t)]
		column = ws['D']
		heightCol = [column[x].value for x in range(t)]
		column = ws['E']
		widthCol = [column[x].value for x in range(t)]
		column = ws['F']
		scaleByCol = [column[x].value for x in range(t)]

		for i in range(1,t):
			x = idCol[i]
			classesDict[x] = [classCol[i], cpriceCol[i], heightCol[i], widthCol[i], scaleByCol[i]]

		return classesDict


	def predictType(self, img):
		#tmp = random.randint(0,self.classesNum)
		#tmp = self.classesNames[tmp]
		#return tmp
		tmp = self.ImgPrepare2(img,self.classesNames[0][2], self.classesNames[0][3])
		prediction = self.model.predict(tmp)  # REMEMBER YOU'RE PASSING A LIST OF THINGS YOU WISH TO PREDICT []
		print(prediction)
		#predictedFruit = classesNames[int(prediction[0][0])]
		m = np.argmax(prediction[0])
		return classesDict[m]


		#return classesDict[int(prediction[0][0])]

	def initCnn(self, modelName):
		print("1:  model is uploaded")
		model = tf.keras.models.load_model(modelName)
		return model

	def ImgPrepare(self,img,r, c):
		#print("Image is converted to ("+str(r) + ", " + str(c)+")")
		#image_array = np.array(image_list)
		new_array = cv2.resize(img, (r,c))  # resize image to match model's expected sizing
		return new_array.reshape(-1, r, c, 1)  # return the image with shaping that TF wants.

	def ImgPrepare1(self, img, r, c):
		#print("Image is converted to (" + str(r) + ", " + str(c) + ")")
		#image_array = np.array(image_list)
		new_array = keras.preprocessing.image.img_to_array(img)
		im = tf.reshape(new_array, [-1, r, c, 3])
		return im

	def ImgPrepare2(self, img, r, c):
		img = cv2.resize(img, (r,c))
		img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
		image_array = np.array(img)

		#print("Image is converted to (" + str(r) + ", " + str(c) + ")")
		#image_array = np.array(image_list)
		#new_array = keras.preprocessing.image.img_to_array(img)
		image_array = tf.reshape(image_array, [-1, r, c, 3])
		return image_array




"""
def predictor(sdir, csv_path,  model_path, averaged=True, verbose=True):    
    # read in the csv file
    class_df=pd.read_csv(csv_path)    
    class_count=len(class_df['class'].unique())
    img_height=int(class_df['height'].iloc[0])
    img_width =int(class_df['width'].iloc[0])
    img_size=(img_width, img_height)    
    scale=class_df['scale by'].iloc[0] 
    image_list = []
    # determine value to scale image pixels by
    try: 
        s=int(scale)
        s2=1
        s1=0
    except:
        split=scale.split('-')
        s1=float(split[1])
        s2=float(split[0].split('*')[1])
    path_list=[]
    paths=os.listdir(sdir)    
    for f in paths:
        path_list.append(os.path.join(sdir,f))
    if verbose:
        print (' Model is being loaded- this will take about 10 seconds')
    model=load_model(model_path)


    image_count=len(path_list) 
    image_list=[]
    file_list=[]
    good_image_count=0
    for i in range (image_count):        
        try:
            img=cv2.imread(path_list[i])
            img=cv2.resize(img, img_size)
            img=cv2.cvtColor(img, cv2.COLOR_BGR2RGB)            
            good_image_count +=1
            img=img*s2 - s1             
            image_list.append(img)
            file_name=os.path.split(path_list[i])[1]
            file_list.append(file_name)
        except:
            if verbose:
                print ( path_list[i], ' is an invalid image file')
    if good_image_count==1: # if only a single image need to expand dimensions
        averaged=True
    image_array=np.array(image_list)    
    # make predictions on images, sum the probabilities of each class then find class index with
    # highest probability
    preds = model.predict(image_array)
    if averaged:
        psum=[]
        for i in range (class_count): # create all 0 values list
            psum.append(0)    
        for p in preds: # iterate over all predictions
            for i in range (class_count):
                psum[i]=psum[i] + p[i]  # sum the probabilities   
        index=np.argmax(psum) # find the class index with the highest probability sum        
        klass=class_df['class'].iloc[index] # get the class name that corresponds to the index
        prob=psum[index]/good_image_count  # get the probability average         
        # to show the correct image run predict again and select first image that has same index
        for img in image_array:  #iterate through the images    
            test_img=np.expand_dims(img, axis=0) # since it is a single image expand dimensions 
            test_index=np.argmax(model.predict(test_img)) # for this image find the class index with highest probability
            if test_index== index: # see if this image has the same index as was selected previously
                if verbose: # show image and print result if verbose=1
                    plt.axis('off')
                    plt.imshow(img) # show the image
                    print (f'predicted species is {klass} with a probability of {prob:6.4f} ')
                break # found an image that represents the predicted class      
        return klass, prob, img, None
    else: # create individual predictions for each image
        pred_class=[]
        prob_list=[]
        for i, p in enumerate(preds):
            index=np.argmax(p) # find the class index with the highest probability sum
            klass=class_df['class'].iloc[index] # get the class name that corresponds to the index
            image_file= file_list[i]
            pred_class.append(klass)
            prob_list.append(p[index])            
        Fseries=pd.Series(file_list, name='image file')
        Lseries=pd.Series(pred_class, name= 'species')
        Pseries=pd.Series(prob_list, name='probability')
        df=pd.concat([Fseries, Lseries, Pseries], axis=1)
        if verbose:
            length= len(df)
            print (df.head(length))
        return None, None, None, df

csv_file='class_dict.csv'
store_path='storage'


csv_path=csv_file # path to class_dict.csv
model_path=model_file # path to the trained model
klass, prob, img, df =predictor(store_path, csv_path,  model_path, averaged=True, verbose=False) # run the classifier
msg=f' image is of  {klass}  with a probability of {prob * 100: 6.2f} %'
plt.axis('off')
plt.imshow(img)
print_in_color(msg, (0,255,255), (65,85,55))




 witherrors
	#data = pd.read_excel(file)
	#idCol = pd.DataFrame(data, columns=["id"])
	#classCol =pd.DataFrame(data, columns=["class"])
	#idCol = data(["id"]).tolist()
	#classCol =data(["class"]).tolist()
	#idCol = list(data(["id"]))
	#classCol =list(data(["class"]))
def processCapturedImg(self,img, predictedFruit):# edit
	cv2.putText(img, predictedFruit, (200,200), cv2.FONT_HERSHEY_PLAIN,2, (255, 0, 255), 2)
	return img
	"""